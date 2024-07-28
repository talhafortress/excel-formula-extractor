import openpyxl

# Load the original Excel file
original_wb = openpyxl.load_workbook('original_file.xlsx', data_only=False)

# Create a new Excel workbook
new_wb = openpyxl.Workbook()
new_wb.remove(new_wb.active)  # Remove the default sheet

# Copy all sheets and preserve formulas
for sheet_name in original_wb.sheetnames:
    original_ws = original_wb[sheet_name]
    new_ws = new_wb.create_sheet(title=sheet_name)
    
    for row in original_ws.iter_rows():
        for cell in row:
            if cell.value and cell.data_type == 'f':  # If the cell has a formula
                new_ws[cell.coordinate].value = cell.value
            else:  # Clear the cell value
                new_ws[cell.coordinate].value = None

# Save the new workbook
new_wb.save('new_file.xlsx')

print("Formulas have been transferred and cell values cleared.")
